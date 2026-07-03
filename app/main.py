from __future__ import annotations

import csv
import base64
import hashlib
import html
import json
import math
import mimetypes
import os
import re
import shutil
import sqlite3
import threading
import uuid
from contextlib import contextmanager
from datetime import datetime, timezone
from collections import deque
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, unquote, urlparse

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from . import accounts, coder

try:
    from llama_cpp import Llama
except Exception:  # pragma: no cover - dependency is optional at boot
    Llama = None  # type: ignore[assignment]

ROOT = Path(__file__).resolve().parents[1]
STATIC_DIR = ROOT / "static"
UPLOAD_DIR = ROOT / "uploads"
DATA_DIR = ROOT / "data"
MODEL_DIR = ROOT / "models" / "Huihui-Qwythos-9B-Claude-Mythos-5-1M-abliterated-GGUF"
MODEL_FILE = MODEL_DIR / "Huihui-Qwythos-9B-Claude-Mythos-5-1M-abliterated-Q4_K.gguf"
DB_PATH = DATA_DIR / "caraxes.db"

APP_NAME = "Caraxes Local AI"
MODEL_REPO = "huihui-ai/Huihui-Qwythos-9B-Claude-Mythos-5-1M-abliterated-GGUF"
MODEL_FILENAME = MODEL_FILE.name
FILE_CHUNK_CHARS = int(os.environ.get("CARAXES_FILE_CHUNK_CHARS", "4500"))
FILE_CHUNK_OVERLAP_CHARS = int(os.environ.get("CARAXES_FILE_CHUNK_OVERLAP_CHARS", "450"))
MAX_TOTAL_CONTEXT_CHARS = int(os.environ.get("CARAXES_PROMPT_FILE_BUDGET_CHARS", "14000"))
MAX_SELECTED_CHUNKS_PER_FILE = int(os.environ.get("CARAXES_MAX_SELECTED_CHUNKS_PER_FILE", "6"))
MAX_HISTORY_MESSAGES = 16
MAX_HISTORY_CONTEXT_CHARS = int(os.environ.get("CARAXES_HISTORY_BUDGET_CHARS", "7000"))
STREAM_TEXT_THRESHOLD_BYTES = int(os.environ.get("CARAXES_STREAM_TEXT_THRESHOLD_BYTES", str(8 * 1024 * 1024)))
MAX_STORED_EXTRACTED_CHARS = int(os.environ.get("CARAXES_MAX_STORED_EXTRACTED_CHARS", "200000"))
MAX_MEMORY_CONTEXT_CHARS = int(os.environ.get("CARAXES_MEMORY_BUDGET_CHARS", "5000"))
MAX_CHAT_MEMORY_CHARS = int(os.environ.get("CARAXES_CHAT_MEMORY_CHARS", "6000"))
WEB_ENABLED = os.environ.get("CARAXES_ENABLE_WEB", "1").lower() not in {"0", "false", "no", "off"}
WEB_MAX_RESULTS = int(os.environ.get("CARAXES_WEB_MAX_RESULTS", "4"))
WEB_FETCH_CHARS = int(os.environ.get("CARAXES_WEB_FETCH_CHARS", "5000"))
WEB_CONTEXT_CHARS = int(os.environ.get("CARAXES_WEB_CONTEXT_CHARS", "12000"))
WEB_TIMEOUT_SECONDS = float(os.environ.get("CARAXES_WEB_TIMEOUT_SECONDS", "12"))
GITHUB_MAX_FILES = int(os.environ.get("CARAXES_GITHUB_MAX_FILES", "8"))
GITHUB_CONTEXT_CHARS = int(os.environ.get("CARAXES_GITHUB_CONTEXT_CHARS", "18000"))
VECTOR_DIMS = int(os.environ.get("CARAXES_VECTOR_DIMS", "384"))
MAX_VECTOR_CHUNKS_PER_FILE = int(os.environ.get("CARAXES_MAX_VECTOR_CHUNKS_PER_FILE", "6"))

SYSTEM_PROMPT = """You are Caraxes, a local coding and research assistant running on the user's machine.
Be direct, helpful, and technically careful. When the user attaches files, use their extracted content as source context.
If file extraction notes say OCR or parsing was unavailable, explain that limitation instead of pretending you saw details.
For code review or suggestions, cite filenames and concrete changes when possible.
Never dump or repeat an attached file verbatim unless the user explicitly asks you to reproduce it.
For cybersecurity, help with authorized defensive work: blue-team triage, purple-team validation, secure coding, detection engineering, incident response, and lab-safe red-team-aware reasoning. Refuse malware, credential theft, phishing kits, stealth, persistence, evasion, destructive actions, or unauthorized intrusion instructions.
When the user asks what code does, explain it before showing code. Use this shape when helpful:
Summary, How I read it, What it does, Important code, Risks or issues, Suggestions.
Use those as headings, not as a literal opening sentence.
For "How I read it", provide concise observable reasoning steps, not private hidden chain-of-thought.
Quote only short code snippets that support the explanation.
Keep any internal reasoning brief and always emit a visible final answer."""

STOP_WORDS = {
    "about",
    "after",
    "again",
    "also",
    "and",
    "are",
    "attached",
    "can",
    "could",
    "does",
    "file",
    "files",
    "for",
    "from",
    "give",
    "have",
    "how",
    "into",
    "its",
    "just",
    "like",
    "make",
    "me",
    "only",
    "please",
    "read",
    "show",
    "tell",
    "that",
    "the",
    "this",
    "what",
    "when",
    "with",
    "you",
}

TEXT_EXTRA_SUFFIXES = {
    ".asm",
    ".s",
    ".nasm",
    ".masm",
    ".dis",
    ".log",
    ".trace",
    ".out",
    ".err",
}

BINARY_SUFFIXES = {
    ".bin",
    ".dat",
    ".exe",
    ".dll",
    ".sys",
    ".so",
    ".dylib",
    ".o",
    ".obj",
    ".class",
    ".jar",
    ".wasm",
    ".pyc",
}

REPO_TEXT_SUFFIXES = {
    ".asm",
    ".bat",
    ".c",
    ".cfg",
    ".cpp",
    ".cs",
    ".css",
    ".go",
    ".h",
    ".hpp",
    ".html",
    ".ini",
    ".java",
    ".js",
    ".json",
    ".kt",
    ".log",
    ".md",
    ".php",
    ".ps1",
    ".py",
    ".rb",
    ".rs",
    ".s",
    ".sh",
    ".sql",
    ".toml",
    ".ts",
    ".tsx",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}

REPO_IMPORTANT_NAMES = {
    "readme.md",
    "readme",
    "pyproject.toml",
    "requirements.txt",
    "package.json",
    "composer.json",
    "go.mod",
    "cargo.toml",
    "pom.xml",
    "build.gradle",
    "dockerfile",
    "docker-compose.yml",
    ".github/workflows",
}

LOG_SIGNAL_TERMS = {
    "error": "error",
    "exception": "exception",
    "traceback": "traceback",
    "warning": "warning",
    "warn": "warning",
    "failed": "failed",
    "failure": "failed",
    "fatal": "fatal",
    "timeout": "timeout",
    "denied": "denied",
    "segmentation": "segmentation",
    "nullpointer": "nullpointer",
    "panic": "panic",
}

TASK_MODE_ALIASES = {
    "auto": "auto",
    "coding": "coding",
    "code": "coding",
    "code_review": "code_review",
    "secure_code": "secure_code",
    "secure_review": "secure_code",
    "blue": "blue_team",
    "blue_team": "blue_team",
    "incident": "incident_response",
    "incident_response": "incident_response",
    "timeline": "incident_response",
    "purple": "purple_team",
    "purple_team": "purple_team",
    "detection": "detection_engineering",
    "detection_engineering": "detection_engineering",
    "red_defense": "red_team_aware_defense",
    "red_team_aware_defense": "red_team_aware_defense",
    "large_file": "large_file",
}

TASK_MODE_INSTRUCTIONS = {
    "coding": """Mode: Coding assistant.
Explain code paths precisely, identify bugs and risky assumptions, suggest the smallest safe change first, and list tests or commands to verify. When editing is requested, prefer complete file patches or exact snippets with filenames.""",
    "code_review": """Mode: Code review.
Lead with severity-ordered findings. For each finding include affected file/function when available, impact, evidence, minimal fix direction, and tests. Keep summaries after findings.""",
    "secure_code": """Mode: Secure code review.
Map issues to CWE/OWASP where useful, explain exploitability at a high level from a defender perspective, avoid weaponized steps, and provide safer implementation patterns plus abuse-case tests.""",
    "blue_team": """Mode: Blue-team triage.
Return executive summary, timeline, suspicious entities, evidence, likely false positives, MITRE ATT&CK mapping, containment steps, detection improvements, and missing evidence. Separate facts from hypotheses.""",
    "incident_response": """Mode: Incident response timeline.
Build a timeline from evidence, list affected users/hosts/IPs/processes/files/domains, assign confidence, identify containment and eradication actions, and name exact missing logs or chunks needed next.""",
    "purple_team": """Mode: Purple-team validation.
Use red-team knowledge at TTP level only. Return technique summary, adversary behavior, lab-safe emulation idea, required telemetry, detection logic ideas, expected alerts, hardening controls, success criteria, and backlog.""",
    "detection_engineering": """Mode: Detection engineering.
Evaluate detections for coverage, false positives, false negatives, telemetry requirements, MITRE mapping, test cases, tuning notes, and safe rollout steps.""",
    "red_team_aware_defense": """Mode: Red-team-aware defense.
Explain attack paths from a defender perspective: prerequisites, attacker decision points, observable telemetry, prevention, detection, containment. Do not provide exploit code, malware, credential theft, stealth, persistence, evasion, or unauthorized intrusion instructions.""",
    "large_file": """Mode: Large-file analysis.
First map the evidence by chunks/parts, then inspect relevant parts, then synthesize. Never pretend omitted chunks were read. Cite included parts and ask for exact next chunk numbers or search terms when coverage is insufficient.""",
}

DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)
MODEL_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title=APP_NAME)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


@contextmanager
def db() -> Any:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            create table if not exists chats (
                id text primary key,
                title text not null,
                created_at text not null,
                updated_at text not null
            );

            create table if not exists messages (
                id text primary key,
                chat_id text not null references chats(id) on delete cascade,
                role text not null,
                content text not null,
                attachments_json text not null default '[]',
                created_at text not null
            );

            create table if not exists files (
                id text primary key,
                original_name text not null,
                stored_name text not null,
                path text not null,
                mime_type text,
                extracted_text text not null,
                meta_json text not null,
                created_at text not null
            );

            create table if not exists file_chunks (
                id text primary key,
                file_id text not null references files(id) on delete cascade,
                chunk_index integer not null,
                total_chunks integer not null,
                text text not null,
                char_count integer not null,
                created_at text not null
            );

            create index if not exists idx_file_chunks_file_id
            on file_chunks (file_id, chunk_index);

            create table if not exists chat_memory (
                chat_id text primary key references chats(id) on delete cascade,
                summary text not null,
                updated_at text not null
            );

            create table if not exists memory_items (
                id text primary key,
                kind text not null,
                content text not null,
                source text,
                created_at text not null,
                updated_at text not null
            );

            create table if not exists file_chunk_vectors (
                file_id text not null references files(id) on delete cascade,
                chunk_index integer not null,
                vector_json text not null,
                updated_at text not null,
                primary key (file_id, chunk_index)
            );
            """
        )
        try:
            conn.execute(
                """
                create virtual table if not exists file_chunks_fts
                using fts5(file_id unindexed, chunk_index unindexed, original_name, text)
                """
            )
        except sqlite3.OperationalError:
            # Some embedded SQLite builds omit FTS5. Retrieval still works through keyword scoring.
            pass


init_db()
accounts.init(DB_PATH)
app.include_router(accounts.router)
app.middleware("http")(accounts.auth_middleware)


class ChatCreate(BaseModel):
    title: str | None = None


class ChatMessageIn(BaseModel):
    content: str = Field(default="", max_length=120_000)
    attachment_ids: list[str] = Field(default_factory=list)
    mode: str = Field(default="auto", max_length=40)
    temperature: float = Field(default=0.7, ge=0.0, le=2.0)
    max_tokens: int = Field(default=900, ge=64, le=4096)


class ChatMessageOut(BaseModel):
    chat_id: str
    user_message: dict[str, Any]
    assistant_message: dict[str, Any]
    model_status: dict[str, Any]
    research_sources: list[dict[str, Any]] = Field(default_factory=list)


def row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    item = dict(row)
    for key in ("attachments_json", "meta_json"):
        if key in item:
            target_key = key.replace("_json", "")
            item[target_key] = json.loads(item.pop(key) or "[]")
    return item


class VisibleTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._skip_depth = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip_depth += 1
        if tag in {"p", "br", "div", "section", "article", "h1", "h2", "h3", "li", "tr"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript", "svg"} and self._skip_depth:
            self._skip_depth -= 1
        if tag in {"p", "div", "section", "article", "li", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._skip_depth:
            text = data.strip()
            if text:
                self.parts.append(text)

    def text(self) -> str:
        joined = " ".join(self.parts)
        joined = re.sub(r"[ \t\r\f\v]+", " ", joined)
        joined = re.sub(r"\n\s*", "\n", joined)
        joined = re.sub(r"\n{3,}", "\n\n", joined)
        return html.unescape(joined).strip()


def compact_text(text: str, limit: int) -> str:
    clean = re.sub(r"\s+", " ", (text or "")).strip()
    if len(clean) <= limit:
        return clean
    return clean[: max(0, limit - 3)].rstrip() + "..."


def clip_text(text: str, limit: int) -> str:
    value = (text or "").strip()
    if len(value) <= limit:
        return value
    return value[: max(0, limit - 3)].rstrip() + "..."


def explicit_memory_notes(text: str) -> list[str]:
    notes: list[str] = []
    for match in re.finditer(r"(?:remember|keep in memory|save this)\s*:?\s*(.+)", text, flags=re.I):
        note = match.group(1).strip()
        if note:
            notes.append(compact_text(note, 800))
    return notes[:5]


def store_explicit_memories(text: str, source: str) -> None:
    notes = explicit_memory_notes(text)
    if not notes:
        return
    ts = now_iso()
    with db() as conn:
        for note in notes:
            conn.execute(
                """
                insert into memory_items (id, kind, content, source, created_at, updated_at)
                values (?, 'user_note', ?, ?, ?, ?)
                """,
                (str(uuid.uuid4()), note, source, ts, ts),
            )


def build_memory_context(chat_id: str) -> str:
    blocks: list[str] = []
    with db() as conn:
        chat_row = conn.execute("select summary from chat_memory where chat_id = ?", (chat_id,)).fetchone()
        memory_rows = conn.execute(
            """
            select kind, content, source, updated_at
            from memory_items
            order by updated_at desc
            limit 12
            """
        ).fetchall()
    if chat_row and chat_row["summary"].strip():
        blocks.append("Current chat working memory:\n" + chat_row["summary"].strip())
    if memory_rows:
        notes = [f"- {row['content']}" for row in memory_rows if row["content"]]
        if notes:
            blocks.append("Long-term user/project memory:\n" + "\n".join(notes))
    if not blocks:
        return ""
    return compact_text(
        "Persistent memory. Use this to continue previous work after context trimming; do not mention it unless relevant.\n\n"
        + "\n\n".join(blocks),
        MAX_MEMORY_CONTEXT_CHARS,
    )


def update_chat_memory(chat_id: str, user_text: str, files: list[dict[str, Any]], assistant_text: str) -> None:
    attachments = ", ".join(
        f"{item['original_name']} ({item.get('meta', {}).get('chunk_count', 0)} parts)"
        for item in files[:8]
    )
    latest = [
        f"Latest user request: {compact_text(user_text or '[Attached files]', 1200)}",
    ]
    if attachments:
        latest.append(f"Latest attached files: {attachments}")
    latest.append(f"Latest assistant result: {compact_text(assistant_text, 1200)}")
    latest_block = "\n".join(latest)

    with db() as conn:
        row = conn.execute("select summary from chat_memory where chat_id = ?", (chat_id,)).fetchone()
        previous = (row["summary"] if row else "").strip()
        summary = compact_text((previous + "\n\n" + latest_block).strip(), MAX_CHAT_MEMORY_CHARS)
        ts = now_iso()
        conn.execute(
            """
            insert into chat_memory (chat_id, summary, updated_at)
            values (?, ?, ?)
            on conflict(chat_id) do update set summary = excluded.summary, updated_at = excluded.updated_at
            """,
            (chat_id, summary, ts),
        )


def user_wants_web(query: str) -> bool:
    if not WEB_ENABLED:
        return False
    lower = query.lower()
    if re.search(r"https?://", query):
        return True
    triggers = {
        "internet",
        "web",
        "browser",
        "browse",
        "search",
        "google",
        "google it",
        "look up",
        "look for",
        "latest",
        "current",
        "today",
        "online",
        "website",
        "source",
        "sources",
        "docs",
        "documentation",
        "download",
        "install",
        "github",
        "huggingface",
    }
    return any(trigger in lower for trigger in triggers)


def public_http_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return False
    host = parsed.hostname or ""
    if host in {"localhost"} or host.endswith(".local"):
        return False
    try:
        import ipaddress

        ip = ipaddress.ip_address(host)
        return not (ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_multicast)
    except ValueError:
        return True


def urls_from_text(text: str) -> list[str]:
    urls = []
    for match in re.finditer(r"https?://[^\s<>\")]+", text):
        url = match.group(0).rstrip(".,;:")
        if public_http_url(url) and url not in urls:
            urls.append(url)
    return urls[:WEB_MAX_RESULTS]


def html_to_visible_text(value: str) -> str:
    parser = VisibleTextParser()
    try:
        parser.feed(value)
        return parser.text()
    except Exception:
        stripped = re.sub(r"<[^>]+>", " ", value)
        return compact_text(html.unescape(stripped), WEB_FETCH_CHARS)


def http_get_text(url: str, headers: dict[str, str] | None = None, timeout: float = WEB_TIMEOUT_SECONDS) -> tuple[str, dict[str, str], str]:
    headers = headers or {}
    try:
        import httpx

        with httpx.Client(timeout=timeout, follow_redirects=True, headers=headers) as client:
            response = client.get(url)
            response.raise_for_status()
            return str(response.url), dict(response.headers), response.text
    except ModuleNotFoundError:
        pass

    import urllib.request

    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=timeout) as response:  # nosec - URLs are public-http filtered by callers.
        raw = response.read()
        content_type = response.headers.get("content-type", "")
        charset_match = re.search(r"charset=([^;\s]+)", content_type, flags=re.I)
        charset = charset_match.group(1) if charset_match else "utf-8"
        text = raw.decode(charset, errors="replace")
        return response.geturl(), dict(response.headers.items()), text


def fetch_web_page(url: str) -> dict[str, str] | None:
    if not public_http_url(url):
        return None
    try:
        final_url, headers, text = http_get_text(
            url,
            headers={"User-Agent": "CaraxesLocalAI/1.0 (+local research assistant)"},
        )
        content_type = headers.get("content-type", "") or headers.get("Content-Type", "")
        text = text if "text" in content_type or "html" in content_type or "json" in content_type or not content_type else ""
        if not text:
            return {"url": final_url, "title": "", "text": f"[Fetched non-text content: {content_type}]"}
        title_match = re.search(r"<title[^>]*>(.*?)</title>", text, flags=re.I | re.S)
        title = html.unescape(re.sub(r"\s+", " ", title_match.group(1)).strip()) if title_match else ""
        visible = html_to_visible_text(text)
        return {"url": final_url, "title": title, "text": visible[:WEB_FETCH_CHARS]}
    except Exception:
        return None


def decode_duckduckgo_url(url: str) -> str:
    if url.startswith("//"):
        url = "https:" + url
    parsed = urlparse(url)
    if "duckduckgo.com" in (parsed.hostname or ""):
        target = parse_qs(parsed.query).get("uddg")
        if target:
            return unquote(target[0])
    return url


def parse_duckduckgo_results(page_text: str) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    pattern = re.compile(
        r"<a(?=[^>]*class=['\"]result-link['\"])[^>]*href=['\"]([^'\"]+)['\"][^>]*>(.*?)</a>"
        r".*?<td[^>]+class=['\"]result-snippet['\"][^>]*>(.*?)</td>",
        flags=re.I | re.S,
    )
    for match in pattern.finditer(page_text):
        url = decode_duckduckgo_url(html.unescape(match.group(1)))
        if not public_http_url(url):
            continue
        title = compact_text(html_to_visible_text(match.group(2)), 180)
        snippet = compact_text(html_to_visible_text(match.group(3)), 500)
        if url and not any(item["url"] == url for item in results):
            results.append({"title": title, "url": url, "snippet": snippet})
        if len(results) >= WEB_MAX_RESULTS:
            break
    return results


def search_web(query: str) -> list[dict[str, str]]:
    search_urls = [
        f"https://lite.duckduckgo.com/lite/?q={quote_plus(query)}",
        f"https://html.duckduckgo.com/html/?q={quote_plus(query)}",
    ]
    for search_url in search_urls:
        try:
            _, _, text = http_get_text(
                search_url,
                headers={
                    "User-Agent": "Mozilla/5.0",
                    "Accept": "text/html,application/xhtml+xml",
                },
            )
        except Exception:
            continue
        results = parse_duckduckgo_results(text)
        if results:
            return results
    return []


WEB_QUERY_DROP_TERMS = {
    "answer",
    "brief",
    "browse",
    "browser",
    "cite",
    "google",
    "internet",
    "online",
    "reply",
    "respond",
    "search",
    "sentence",
    "short",
    "source",
    "sources",
    "url",
    "urls",
    "web",
    "website",
}


def clean_web_search_query(query: str) -> str:
    text = re.sub(r"https?://\S+", " ", query or "")
    text = re.sub(
        r"(?is)\b(answer|reply|respond)\b\s+(in|with|using|as).*$",
        " ",
        text,
    )
    text = re.sub(r"(?is)\b(cite|include|show)\b\s+(the\s+)?(source|sources|url|urls).*$", " ", text)
    terms = []
    for term in query_terms(text):
        clean = term.strip(".,;:!?()[]{}\"'")
        if len(clean) >= 3 and clean not in WEB_QUERY_DROP_TERMS:
            terms.append(clean)
    if terms:
        return " ".join(terms[:16])
    return compact_text(text, 240)


def build_web_context(query: str, source_sink: list[dict[str, Any]] | None = None) -> str:
    if not user_wants_web(query):
        return ""

    urls = urls_from_text(query)
    results = [{"title": "", "url": url, "snippet": "URL provided by the user."} for url in urls]
    if not results:
        raw_query = re.sub(r"https?://\S+", "", query).strip()
        search_queries = []
        for candidate in (clean_web_search_query(raw_query), raw_query):
            candidate = candidate.strip()
            if candidate and candidate not in search_queries:
                search_queries.append(candidate)
        for search_query in search_queries:
            results = search_web(search_query[:500])
            if results:
                break

    blocks: list[str] = []
    fetched = 0
    for index, result in enumerate(results[:WEB_MAX_RESULTS], start=1):
        title = result.get("title") or "Untitled"
        url = result.get("url") or ""
        snippet = result.get("snippet") or ""
        block = [f"### Web source {index}: {title}", f"URL: {url}"]
        if snippet:
            block.append(f"Search snippet: {snippet}")
        page = fetch_web_page(url) if fetched < 2 else None
        if page and page.get("text"):
            fetched += 1
            if page.get("title"):
                block[0] = f"### Web source {index}: {page['title']}"
            block.append("Fetched text:\n" + page["text"][:WEB_FETCH_CHARS])
        if source_sink is not None:
            source_sink.append(
                {
                    "kind": "web",
                    "name": page.get("title") if page and page.get("title") else title,
                    "url": page.get("url") if page and page.get("url") else url,
                    "snippet": snippet,
                    "fetched": bool(page and page.get("text")),
                }
            )
        blocks.append("\n".join(block))

    if not blocks:
        return "Web access was requested, but the free search/fetch path did not return readable results."
    return clip_text(
        "Web research context. Treat these snippets as fresh but imperfect external sources. Cite URLs in the answer when used.\n\n"
        + "\n\n".join(blocks),
        WEB_CONTEXT_CHARS,
    )


def github_repos_from_text(text: str) -> list[tuple[str, str]]:
    repos: list[tuple[str, str]] = []
    for match in re.finditer(r"https?://github\.com/([A-Za-z0-9_.-]+)/([A-Za-z0-9_.-]+)", text):
        owner = match.group(1)
        repo = match.group(2).removesuffix(".git")
        if (owner, repo) not in repos:
            repos.append((owner, repo))
    return repos[:3]


def github_headers() -> dict[str, str]:
    headers = {"Accept": "application/vnd.github+json", "User-Agent": "CaraxesLocalAI/1.0"}
    token = os.environ.get("GITHUB_TOKEN") or os.environ.get("CARAXES_GITHUB_TOKEN")
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def repo_file_score(path: str, query_terms_list: list[str]) -> int:
    lower = path.lower()
    name = Path(lower).name
    score = 0
    if name in REPO_IMPORTANT_NAMES or lower in REPO_IMPORTANT_NAMES:
        score += 40
    if lower.startswith(("docs/", "doc/", ".github/workflows/", "src/", "app/", "lib/", "scripts/")):
        score += 10
    if Path(lower).suffix in {".py", ".js", ".ts", ".java", ".c", ".cpp", ".go", ".rs", ".ps1", ".sh", ".md"}:
        score += 8
    for term in query_terms_list:
        if term in lower:
            score += 12
    return score


def github_api_json(client: Any, url: str) -> Any | None:
    try:
        response = client.get(url)
        if response.status_code == 403:
            return {"error": "GitHub API rate limit or permission denied."}
        response.raise_for_status()
        return response.json()
    except Exception:
        return None


def fetch_github_file(client: Any, owner: str, repo: str, branch: str, path: str, max_chars: int) -> str:
    if not path or Path(path).suffix.lower() not in REPO_TEXT_SUFFIXES and Path(path).name.lower() not in REPO_IMPORTANT_NAMES:
        return ""
    raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{path}"
    try:
        response = client.get(raw_url)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "")
        if "text" not in content_type and "json" not in content_type and "xml" not in content_type and content_type:
            return ""
        return response.text[:max_chars]
    except Exception:
        return ""


def fetch_github_readme(client: Any, owner: str, repo: str, branch: str) -> str:
    api_url = f"https://api.github.com/repos/{owner}/{repo}/readme"
    data = github_api_json(client, api_url)
    if isinstance(data, dict) and data.get("content"):
        try:
            decoded = base64.b64decode(data["content"]).decode("utf-8", errors="replace")
            return decoded[:WEB_FETCH_CHARS]
        except Exception:
            pass
    for candidate in ("README.md", "README.rst", "README.txt", "README"):
        text = fetch_github_file(client, owner, repo, branch, candidate, WEB_FETCH_CHARS)
        if text:
            return text
    return ""


def build_github_repo_context(owner: str, repo: str, query: str) -> str:
    try:
        import httpx

        with httpx.Client(timeout=WEB_TIMEOUT_SECONDS, headers=github_headers(), follow_redirects=True) as client:
            repo_data = github_api_json(client, f"https://api.github.com/repos/{owner}/{repo}")
            if not isinstance(repo_data, dict) or repo_data.get("error"):
                return f"### GitHub repository: {owner}/{repo}\nCould not fetch repository metadata."
            branch = repo_data.get("default_branch") or "main"
            readme = fetch_github_readme(client, owner, repo, branch)
            tree_data = github_api_json(
                client,
                f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1",
            )
            tree = tree_data.get("tree", []) if isinstance(tree_data, dict) else []
            terms = query_terms(query)
            files = [
                item
                for item in tree
                if item.get("type") == "blob"
                and int(item.get("size") or 0) <= 250_000
                and (
                    Path(item.get("path", "")).suffix.lower() in REPO_TEXT_SUFFIXES
                    or Path(item.get("path", "")).name.lower() in REPO_IMPORTANT_NAMES
                )
            ]
            files.sort(key=lambda item: repo_file_score(item.get("path", ""), terms), reverse=True)
            selected = files[:GITHUB_MAX_FILES]
            file_blocks: list[str] = []
            per_file_chars = max(1500, GITHUB_CONTEXT_CHARS // max(1, len(selected) + 2))
            for item in selected:
                path = item.get("path", "")
                text = fetch_github_file(client, owner, repo, branch, path, per_file_chars)
                if text:
                    file_blocks.append(f"--- Repo file: {path} ---\n{text[:per_file_chars]}")
    except Exception as exc:
        return f"### GitHub repository: {owner}/{repo}\nGitHub fetch failed: {exc}"

    summary = [
        f"### GitHub repository: {owner}/{repo}",
        f"URL: https://github.com/{owner}/{repo}",
        f"Description: {repo_data.get('description') or 'No description'}",
        f"Default branch: {branch}",
        f"Stars: {repo_data.get('stargazers_count')}, forks: {repo_data.get('forks_count')}",
        f"Language: {repo_data.get('language') or 'unknown'}",
        f"Selected files: {', '.join(item.get('path', '') for item in selected) if selected else 'none'}",
    ]
    if readme:
        summary.append("--- README preview ---\n" + readme[:WEB_FETCH_CHARS])
    summary.extend(file_blocks)
    return clip_text("\n".join(summary), GITHUB_CONTEXT_CHARS)


def build_github_context(query: str, source_sink: list[dict[str, Any]] | None = None) -> str:
    repos = github_repos_from_text(query)
    if not repos:
        return ""
    blocks = [
        "GitHub repository context. This is a read-only pull of metadata, README, file tree, and selected files. "
        "Do not claim to have cloned, executed, installed, or audited files that are not included below."
    ]
    for owner, repo in repos:
        if source_sink is not None:
            source_sink.append({"kind": "github", "name": f"{owner}/{repo}", "url": f"https://github.com/{owner}/{repo}"})
        blocks.append(build_github_repo_context(owner, repo, query))
    return clip_text("\n\n".join(blocks), GITHUB_CONTEXT_CHARS * len(repos))


def get_chat_or_404(chat_id: str, user_id: str | None = None) -> dict[str, Any]:
    with db() as conn:
        row = conn.execute("select * from chats where id = ?", (chat_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Chat not found")
    chat = row_to_dict(row)
    if user_id is not None and chat.get("user_id") not in (None, user_id):
        raise HTTPException(status_code=404, detail="Chat not found")
    return chat


def create_chat(title: str | None = None, user_id: str | None = None) -> dict[str, Any]:
    chat_id = str(uuid.uuid4())
    ts = now_iso()
    clean_title = (title or "New chat").strip()[:80] or "New chat"
    with db() as conn:
        conn.execute(
            "insert into chats (id, title, created_at, updated_at, user_id) values (?, ?, ?, ?, ?)",
            (chat_id, clean_title, ts, ts, user_id),
        )
    return {"id": chat_id, "title": clean_title, "created_at": ts, "updated_at": ts, "user_id": user_id}


def insert_message(
    chat_id: str,
    role: str,
    content: str,
    attachments: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    message_id = str(uuid.uuid4())
    ts = now_iso()
    attachments_json = json.dumps(attachments or [], ensure_ascii=False)
    with db() as conn:
        conn.execute(
            """
            insert into messages (id, chat_id, role, content, attachments_json, created_at)
            values (?, ?, ?, ?, ?, ?)
            """,
            (message_id, chat_id, role, content, attachments_json, ts),
        )
        conn.execute("update chats set updated_at = ? where id = ?", (ts, chat_id))
        if role == "user":
            maybe_title = content.strip().replace("\n", " ")
            if maybe_title:
                old = conn.execute("select title from chats where id = ?", (chat_id,)).fetchone()
                if old and old["title"] == "New chat":
                    conn.execute(
                        "update chats set title = ? where id = ?",
                        (maybe_title[:64], chat_id),
                    )
    return {
        "id": message_id,
        "chat_id": chat_id,
        "role": role,
        "content": content,
        "attachments": attachments or [],
        "created_at": ts,
    }


def guess_text_file(path: Path, mime_type: str | None) -> bool:
    if mime_type and (mime_type.startswith("text/") or mime_type in {"application/json", "application/xml"}):
        return True
    return path.suffix.lower() in {
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".jsx",
        ".html",
        ".css",
        ".scss",
        ".json",
        ".jsonl",
        ".md",
        ".txt",
        ".yml",
        ".yaml",
        ".toml",
        ".ini",
        ".cfg",
        ".csv",
        ".sql",
        ".ps1",
        ".bat",
        ".sh",
        ".xml",
        ".java",
        ".c",
        ".cpp",
        ".h",
        ".hpp",
        ".cs",
        ".go",
        ".rs",
        ".php",
        ".rb",
        ".swift",
        ".kt",
    } | TEXT_EXTRA_SUFFIXES


def should_stream_text_file(path: Path, mime_type: str | None) -> bool:
    try:
        size = path.stat().st_size
    except OSError:
        return False
    return size >= STREAM_TEXT_THRESHOLD_BYTES and guess_text_file(path, mime_type)


def hexdump(data: bytes, width: int = 16) -> str:
    lines: list[str] = []
    for offset in range(0, len(data), width):
        chunk = data[offset : offset + width]
        hex_part = " ".join(f"{byte:02x}" for byte in chunk)
        ascii_part = "".join(chr(byte) if 32 <= byte < 127 else "." for byte in chunk)
        lines.append(f"{offset:08x}  {hex_part:<{width * 3}}  {ascii_part}")
    return "\n".join(lines)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def emit_stream_chunk(
    chunks: list[str],
    buffer: list[tuple[int, str]],
    start_line: int,
    end_line: int,
) -> tuple[list[tuple[int, str]], int]:
    if not buffer:
        return buffer, start_line
    text = "".join(line for _, line in buffer).strip()
    if text:
        chunks.append(f"--- Lines {start_line}-{end_line} ---\n{text}")

    overlap: list[tuple[int, str]] = []
    overlap_chars = 0
    for line_no, line in reversed(buffer):
        if overlap_chars >= FILE_CHUNK_OVERLAP_CHARS:
            break
        overlap.append((line_no, line))
        overlap_chars += len(line)
    overlap.reverse()
    next_start = overlap[0][0] if overlap else end_line + 1
    return overlap, next_start


def stream_text_file(path: Path) -> tuple[str, dict[str, Any], list[str]]:
    meta: dict[str, Any] = {
        "kind": "text",
        "notes": ["Large text file streamed into chunks; the stored preview is not the full file."],
        "size_bytes": path.stat().st_size,
        "streamed": True,
    }
    chunks: list[str] = []
    buffer: list[tuple[int, str]] = []
    buffer_chars = 0
    chunk_start_line = 1
    line_count = 0
    char_count = 0
    first_lines: list[str] = []
    last_lines: deque[str] = deque(maxlen=20)
    signal_counts = {label: 0 for label in sorted(set(LOG_SIGNAL_TERMS.values()))}

    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        for line in handle:
            line_count += 1
            char_count += len(line)
            lower = line.lower()
            for needle, label in LOG_SIGNAL_TERMS.items():
                if needle in lower:
                    signal_counts[label] += 1
            if len(first_lines) < 20:
                first_lines.append(line.rstrip("\n\r"))
            last_lines.append(line.rstrip("\n\r"))
            buffer.append((line_count, line))
            buffer_chars += len(line)
            if buffer_chars >= FILE_CHUNK_CHARS:
                buffer, chunk_start_line = emit_stream_chunk(chunks, buffer, chunk_start_line, line_count)
                buffer_chars = sum(len(line_text) for _, line_text in buffer)

    if buffer:
        emit_stream_chunk(chunks, buffer, chunk_start_line, line_count)

    meta["line_count"] = line_count
    meta["char_count"] = char_count
    meta["signal_counts"] = {key: value for key, value in signal_counts.items() if value}
    meta["first_lines"] = first_lines[:10]
    meta["last_lines"] = list(last_lines)[-10:]
    preview_parts = [
        f"[Large text preview: {line_count} lines, {char_count} characters, {len(chunks)} parts.]",
        "--- First lines ---",
        "\n".join(first_lines[:20]),
    ]
    if last_lines:
        preview_parts.extend(["--- Last lines ---", "\n".join(list(last_lines)[-20:])])
    return "\n".join(preview_parts).strip(), meta, chunks


def split_text_chunks(text: str) -> list[str]:
    clean = text.strip()
    if not clean:
        return []
    if len(clean) <= FILE_CHUNK_CHARS:
        return [clean]

    chunks: list[str] = []
    start = 0
    text_len = len(clean)
    while start < text_len:
        target_end = min(text_len, start + FILE_CHUNK_CHARS)
        end = target_end
        if target_end < text_len:
            search_start = max(start + FILE_CHUNK_CHARS // 2, target_end - 900)
            newline = clean.rfind("\n", search_start, target_end)
            if newline > start:
                end = newline

        chunk = clean[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= text_len:
            break
        start = max(end - FILE_CHUNK_OVERLAP_CHARS, start + 1)
    return chunks


def query_terms(query: str) -> list[str]:
    terms = re.findall(r"[a-zA-Z0-9_.$#-]{3,}", query.lower())
    return [term for term in terms if term not in STOP_WORDS]


def vector_terms(text: str) -> list[str]:
    terms = query_terms(text)
    expanded: list[str] = []
    for term in terms[:1200]:
        expanded.append(term)
        if len(term) > 5:
            expanded.append(term[:5])
            expanded.append(term[-5:])
    return expanded


def hash_dim(term: str) -> str:
    digest = hashlib.blake2b(term.encode("utf-8", errors="ignore"), digest_size=4).digest()
    return str(int.from_bytes(digest, "little") % VECTOR_DIMS)


def sparse_vector(text: str) -> dict[str, float]:
    weights: dict[str, float] = {}
    for term in vector_terms(text):
        key = hash_dim(term)
        weights[key] = weights.get(key, 0.0) + 1.0
    if not weights:
        return {}
    norm = math.sqrt(sum(value * value for value in weights.values())) or 1.0
    return {key: round(value / norm, 6) for key, value in weights.items()}


def cosine_sparse(left: dict[str, float], right: dict[str, float]) -> float:
    if not left or not right:
        return 0.0
    if len(left) > len(right):
        left, right = right, left
    return sum(value * right.get(key, 0.0) for key, value in left.items())


def vector_json(text: str) -> str:
    return json.dumps(sparse_vector(text), separators=(",", ":"))


def explicit_chunk_numbers(query: str) -> set[int]:
    selected: set[int] = set()
    lower = query.lower()
    for match in re.finditer(r"(?:part|parts|chunk|chunks)\s+(\d+)(?:\s*[-to]+\s*(\d+))?", lower):
        start = int(match.group(1))
        end = int(match.group(2) or start)
        if end < start:
            start, end = end, start
        selected.update(range(start - 1, min(end, start + 10)))
    return {index for index in selected if index >= 0}


def chunk_score(text: str, terms: list[str], filename: str) -> int:
    haystack = f"{filename}\n{text}".lower()
    score = 0
    for term in terms:
        count = haystack.count(term)
        if count:
            score += count * (4 if term in filename.lower() else 2)
    return score


def log_signal_score(text: str) -> int:
    lower = text.lower()
    score = 0
    for needle, label in LOG_SIGNAL_TERMS.items():
        count = lower.count(needle)
        if count:
            score += count * (4 if label in {"fatal", "error", "exception", "panic"} else 2)
    for needle in ("powershell", "encodedcommand", "mimikatz", "lateral", "rundll32", "regsvr32", "failed login", "admin"):
        score += lower.count(needle) * 3
    return score


def high_signal_chunk_indices(chunks: list[dict[str, Any]], limit: int) -> list[int]:
    scored = [(log_signal_score(chunk["text"]), chunk["chunk_index"]) for chunk in chunks]
    scored = [item for item in scored if item[0] > 0]
    scored.sort(reverse=True)
    return [index for _, index in scored[:limit]]


def fts_query_from_terms(terms: list[str]) -> str:
    phrases: list[str] = []
    for term in terms[:16]:
        cleaned = re.sub(r"[^a-zA-Z0-9_.$#-]+", " ", term).strip()
        if cleaned:
            phrases.append('"' + cleaned.replace('"', '""') + '"')
    return " OR ".join(phrases)


def fts_top_chunk_indices(file_id: str, terms: list[str], limit: int) -> list[int]:
    query = fts_query_from_terms(terms)
    if not query:
        return []
    try:
        with db() as conn:
            rows = conn.execute(
                """
                select chunk_index
                from file_chunks_fts
                where file_chunks_fts match ? and file_id = ?
                order by bm25(file_chunks_fts)
                limit ?
                """,
                (query, file_id, limit),
            ).fetchall()
    except sqlite3.OperationalError:
        return []
    return [int(row["chunk_index"]) for row in rows]


def vector_top_chunk_indices(file_id: str, query: str, limit: int) -> list[int]:
    query_vector = sparse_vector(query)
    if not query_vector:
        return []
    try:
        with db() as conn:
            rows = conn.execute(
                """
                select chunk_index, vector_json
                from file_chunk_vectors
                where file_id = ?
                """,
                (file_id,),
            ).fetchall()
    except sqlite3.OperationalError:
        return []
    scored: list[tuple[float, int]] = []
    for row in rows:
        try:
            chunk_vector = json.loads(row["vector_json"] or "{}")
        except json.JSONDecodeError:
            continue
        score = cosine_sparse(query_vector, chunk_vector)
        if score > 0:
            scored.append((score, int(row["chunk_index"])))
    scored.sort(reverse=True)
    return [index for _, index in scored[:limit]]


def backfill_file_chunk_fts() -> None:
    try:
        with db() as conn:
            conn.execute(
                """
                insert into file_chunks_fts (file_id, chunk_index, original_name, text)
                select fc.file_id, fc.chunk_index, f.original_name, fc.text
                from file_chunks fc
                join files f on f.id = fc.file_id
                where not exists (
                    select 1
                    from file_chunks_fts existing
                    where existing.file_id = fc.file_id
                      and existing.chunk_index = fc.chunk_index
                )
                """
            )
    except sqlite3.OperationalError:
        pass


def backfill_file_chunk_vectors() -> None:
    try:
        with db() as conn:
            rows = conn.execute(
                """
                select fc.file_id, fc.chunk_index, fc.text
                from file_chunks fc
                left join file_chunk_vectors fv
                  on fv.file_id = fc.file_id and fv.chunk_index = fc.chunk_index
                where fv.file_id is null
                """
            ).fetchall()
            ts = now_iso()
            for row in rows:
                conn.execute(
                    """
                    insert or replace into file_chunk_vectors (file_id, chunk_index, vector_json, updated_at)
                    values (?, ?, ?, ?)
                    """,
                    (row["file_id"], row["chunk_index"], vector_json(row["text"]), ts),
                )
    except sqlite3.OperationalError:
        pass


def extract_text(path: Path, mime_type: str | None) -> tuple[str, dict[str, Any]]:
    suffix = path.suffix.lower()
    meta: dict[str, Any] = {"kind": "unknown", "notes": []}

    if suffix == ".pdf" or mime_type == "application/pdf":
        meta["kind"] = "pdf"
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            pages: list[str] = []
            for index, page in enumerate(reader.pages, start=1):
                text = page.extract_text() or ""
                pages.append(f"--- Page {index} ---\n{text.strip()}")
            meta["pages"] = len(reader.pages)
            return "\n\n".join(pages).strip() or "[PDF parsed, but no selectable text was found.]", meta
        except Exception as exc:
            meta["notes"].append(f"PDF extraction failed: {exc}")
            return "[PDF extraction failed.]", meta

    if suffix in {".xlsx", ".xlsm"}:
        meta["kind"] = "spreadsheet"
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in workbook.worksheets:
                lines.append(f"--- Sheet: {sheet.title} ---")
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        lines.append("\t".join(values))
            meta["sheets"] = workbook.sheetnames
            return "\n".join(lines).strip() or "[Spreadsheet parsed, but no cell values were found.]", meta
        except Exception as exc:
            meta["notes"].append(f"Spreadsheet extraction failed: {exc}")
            return "[Spreadsheet extraction failed.]", meta

    if suffix == ".csv":
        meta["kind"] = "csv"
        try:
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                reader = csv.reader(handle)
                return "\n".join("\t".join(row) for row in reader), meta
        except Exception as exc:
            meta["notes"].append(f"CSV extraction failed: {exc}")
            return "[CSV extraction failed.]", meta

    if mime_type and mime_type.startswith("image/"):
        meta["kind"] = "image"
        try:
            from PIL import Image

            with Image.open(path) as image:
                meta.update(
                    {
                        "format": image.format,
                        "width": image.width,
                        "height": image.height,
                        "mode": image.mode,
                    }
                )
                text_parts = [
                    f"Image metadata: format={image.format}, size={image.width}x{image.height}, mode={image.mode}."
                ]
                try:
                    import pytesseract

                    ocr = pytesseract.image_to_string(image).strip()
                    if ocr:
                        text_parts.append("--- OCR text ---\n" + ocr)
                    else:
                        text_parts.append("[OCR ran, but no text was detected in the image.]")
                except Exception as exc:
                    meta["notes"].append(
                        "OCR unavailable. Install the Tesseract application and keep pytesseract installed to read image text."
                    )
                    meta["notes"].append(str(exc))
                    text_parts.append("[Image OCR is not available on this machine yet.]")
                return "\n".join(text_parts), meta
        except Exception as exc:
            meta["notes"].append(f"Image extraction failed: {exc}")
            return "[Image extraction failed.]", meta

    if guess_text_file(path, mime_type):
        meta["kind"] = "text"
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            return text, meta
        except Exception as exc:
            meta["notes"].append(f"Text extraction failed: {exc}")
            return "[Text extraction failed.]", meta

    if suffix in BINARY_SUFFIXES or mime_type == "application/octet-stream":
        meta["kind"] = "binary"
        try:
            size = path.stat().st_size
            with path.open("rb") as handle:
                sample = handle.read(4096)
            meta["size_bytes"] = size
            meta["sha256"] = sha256_file(path)
            return (
                f"Binary file metadata: size={size} bytes, sha256={meta['sha256']}.\n"
                "--- First 4096 bytes as hex ---\n"
                f"{hexdump(sample)}"
            ), meta
        except Exception as exc:
            meta["notes"].append(f"Binary extraction failed: {exc}")
            return "[Binary extraction failed.]", meta

    meta["notes"].append("Unsupported file type for text extraction.")
    return "[Unsupported file type.]", meta


def save_uploaded_file(upload: UploadFile) -> dict[str, Any]:
    original = Path(upload.filename or "upload.bin").name
    stored_name = f"{uuid.uuid4()}-{original}"
    path = UPLOAD_DIR / stored_name
    with path.open("wb") as handle:
        shutil.copyfileobj(upload.file, handle)

    mime_type = upload.content_type or mimetypes.guess_type(original)[0]
    if should_stream_text_file(path, mime_type):
        extracted_text, meta, chunks = stream_text_file(path)
    else:
        extracted_text, meta = extract_text(path, mime_type)
        chunks = split_text_chunks(extracted_text)
    meta.setdefault("char_count", len(extracted_text))
    meta["stored_preview_chars"] = len(extracted_text)
    meta["chunk_count"] = len(chunks)
    stored_extracted_text = extracted_text
    if len(stored_extracted_text) > MAX_STORED_EXTRACTED_CHARS:
        stored_extracted_text = (
            extracted_text[:MAX_STORED_EXTRACTED_CHARS]
            + "\n\n[Full extracted text is stored in numbered chunks for retrieval.]"
        )
    file_id = str(uuid.uuid4())
    created_at = now_iso()
    with db() as conn:
        conn.execute(
            """
            insert into files (id, original_name, stored_name, path, mime_type, extracted_text, meta_json, created_at)
            values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                file_id,
                original,
                stored_name,
                str(path),
                mime_type,
                stored_extracted_text,
                json.dumps(meta, ensure_ascii=False),
                created_at,
            ),
        )
        for index, chunk in enumerate(chunks):
            conn.execute(
                """
                insert into file_chunks (id, file_id, chunk_index, total_chunks, text, char_count, created_at)
                values (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    file_id,
                    index,
                    len(chunks),
                    chunk,
                    len(chunk),
                    created_at,
                ),
            )
            try:
                conn.execute(
                    """
                    insert into file_chunks_fts (file_id, chunk_index, original_name, text)
                    values (?, ?, ?, ?)
                    """,
                    (file_id, index, original, chunk),
                )
            except sqlite3.OperationalError:
                pass
            conn.execute(
                """
                insert or replace into file_chunk_vectors (file_id, chunk_index, vector_json, updated_at)
                values (?, ?, ?, ?)
                """,
                (file_id, index, vector_json(chunk), created_at),
            )
    return {
        "id": file_id,
        "original_name": original,
        "stored_name": stored_name,
        "mime_type": mime_type,
        "extracted_chars": meta.get("char_count", len(extracted_text)),
        "preview": extracted_text[:1200],
        "meta": meta,
        "created_at": created_at,
    }


def fetch_files(file_ids: list[str]) -> list[dict[str, Any]]:
    if not file_ids:
        return []
    placeholders = ",".join("?" for _ in file_ids)
    with db() as conn:
        rows = conn.execute(
            f"select * from files where id in ({placeholders})",
            tuple(file_ids),
        ).fetchall()
    found = {row["id"]: row_to_dict(row) for row in rows}
    return [found[file_id] for file_id in file_ids if file_id in found]


def get_chunks_for_file(item: dict[str, Any]) -> list[dict[str, Any]]:
    with db() as conn:
        rows = conn.execute(
            """
            select chunk_index, total_chunks, text, char_count
            from file_chunks
            where file_id = ?
            order by chunk_index asc
            """,
            (item["id"],),
        ).fetchall()
    if rows:
        return [dict(row) for row in rows]

    transient_chunks = split_text_chunks(item.get("extracted_text") or "")
    total = len(transient_chunks)
    return [
        {
            "chunk_index": index,
            "total_chunks": total,
            "text": chunk,
            "char_count": len(chunk),
        }
        for index, chunk in enumerate(transient_chunks)
    ]


def choose_file_chunks(item: dict[str, Any], query: str, remaining: int) -> list[dict[str, Any]]:
    chunks = get_chunks_for_file(item)
    if not chunks or remaining <= 0:
        return []
    total = len(chunks)
    if total == 1 and chunks[0]["char_count"] <= remaining:
        return chunks

    terms = query_terms(query)
    explicit = explicit_chunk_numbers(query)
    selected: set[int] = set(index for index in explicit if index < total)
    selected.add(0)
    if total > 1:
        selected.add(total - 1)
    if total > 2:
        selected.add(total // 2)

    for index in fts_top_chunk_indices(item["id"], terms, MAX_SELECTED_CHUNKS_PER_FILE):
        if index < total:
            selected.add(index)
        if len(selected) >= MAX_SELECTED_CHUNKS_PER_FILE:
            break

    for index in vector_top_chunk_indices(item["id"], query, MAX_VECTOR_CHUNKS_PER_FILE):
        if index < total:
            selected.add(index)
        if len(selected) >= MAX_SELECTED_CHUNKS_PER_FILE:
            break

    lower_query = query.lower()
    if any(word in lower_query for word in ("log", "incident", "siem", "edr", "alert", "blue", "timeline", "attack")):
        for index in high_signal_chunk_indices(chunks, MAX_SELECTED_CHUNKS_PER_FILE):
            if index < total:
                selected.add(index)
            if len(selected) >= MAX_SELECTED_CHUNKS_PER_FILE:
                break

    scored = sorted(
        (
            (chunk_score(chunk["text"], terms, item["original_name"]), chunk["chunk_index"])
            for chunk in chunks
        ),
        reverse=True,
    )
    for score, index in scored:
        if score <= 0 and len(selected) >= 2:
            continue
        selected.add(index)
        if len(selected) >= MAX_SELECTED_CHUNKS_PER_FILE:
            break

    ordered = [chunk for chunk in chunks if chunk["chunk_index"] in selected]
    fitted: list[dict[str, Any]] = []
    used = 0
    for chunk in ordered:
        if used >= remaining:
            break
        fitted.append(chunk)
        used += chunk["char_count"]
    return fitted


def file_map_summary(item: dict[str, Any], chunks: list[dict[str, Any]], selected_chunks: list[dict[str, Any]]) -> str:
    meta = item.get("meta", {})
    total = len(chunks)
    if total <= 1:
        return ""
    selected = {chunk["chunk_index"] for chunk in selected_chunks}
    high_signal = high_signal_chunk_indices(chunks, 5)
    sampled = sorted({0, total // 4, total // 2, (total * 3) // 4, total - 1})
    lines = [
        "File map:",
        f"- Total parts: {total}. Included in this prompt: {', '.join(str(index + 1) for index in sorted(selected))}.",
        f"- Structural sample parts: {', '.join(str(index + 1) for index in sampled if 0 <= index < total)}.",
    ]
    if high_signal:
        lines.append(f"- High-signal parts by local scan: {', '.join(str(index + 1) for index in high_signal)}.")
    if meta.get("first_lines"):
        lines.append("- First lines preview: " + " | ".join(str(line) for line in meta["first_lines"][:3]))
    if meta.get("last_lines"):
        lines.append("- Last lines preview: " + " | ".join(str(line) for line in meta["last_lines"][-3:]))
    lines.append("- If the answer needs broader coverage, ask to inspect exact omitted part numbers or search terms.")
    return "\n".join(lines)


def build_file_context(files: list[dict[str, Any]], query: str) -> str:
    blocks: list[str] = []
    remaining = MAX_TOTAL_CONTEXT_CHARS
    for item in files:
        if remaining <= 0:
            break
        file_chunks = get_chunks_for_file(item)
        selected_chunks = choose_file_chunks(item, query, remaining)
        if not selected_chunks:
            continue
        notes = item.get("meta", {}).get("notes") or []
        meta = item.get("meta", {})
        header = (
            f"### File: {item['original_name']} ({item.get('mime_type') or 'unknown type'})\n"
            f"Extracted characters: {meta.get('char_count', len(item.get('extracted_text') or ''))}. "
            f"Parts: {len(file_chunks)}. Included parts: "
            f"{', '.join(str(chunk['chunk_index'] + 1) for chunk in selected_chunks)}."
        )
        if meta.get("line_count"):
            header += f"\nLines: {meta['line_count']}."
        if meta.get("size_bytes"):
            header += f"\nSize: {meta['size_bytes']} bytes."
        if meta.get("signal_counts"):
            signals = ", ".join(f"{key}={value}" for key, value in meta["signal_counts"].items())
            header += f"\nDetected log/code signals across the whole file: {signals}."
        if notes:
            header += "\nExtraction notes: " + " | ".join(str(note) for note in notes[:3])
        map_summary = file_map_summary(item, file_chunks, selected_chunks)
        part_blocks = [header]
        if map_summary:
            part_blocks.append(map_summary)
        for chunk in selected_chunks:
            text = chunk["text"]
            allowance = max(0, remaining - 500)
            if allowance <= 0:
                break
            clipped = text[:allowance]
            part_blocks.append(
                f"--- Part {chunk['chunk_index'] + 1} of {chunk['total_chunks']} ---\n{clipped}"
            )
            remaining -= len(clipped)
        blocks.append("\n".join(part_blocks))
    if not blocks:
        return ""
    return (
        "The user attached files. Large files are split into numbered parts, and only the most relevant parts "
        "fit in the current context window. Use the included parts as evidence. If the request needs omitted "
        "parts, say which part numbers should be inspected next instead of pretending you saw everything. "
        "Do not copy these parts back as your answer; answer the user's request in your own explanation and "
        "quote only small snippets when useful.\n\n"
        + "\n\n".join(blocks)
    )


def recent_messages(chat_id: str) -> list[dict[str, Any]]:
    with db() as conn:
        rows = conn.execute(
            """
            select role, content, attachments_json, created_at
            from messages
            where chat_id = ?
            order by created_at desc
            limit ?
            """,
            (chat_id, MAX_HISTORY_MESSAGES),
        ).fetchall()
    items = [row_to_dict(row) for row in rows]
    items.reverse()
    return items


class ModelRuntime:
    def __init__(self) -> None:
        self._llm: Any | None = None
        self._lock = threading.Lock()
        self._chat_lock = threading.Lock()
        self.last_error: str | None = None

    def _context_char_budget(self, max_tokens: int) -> int:
        n_ctx = int(os.environ.get("CARAXES_N_CTX", "8192"))
        prompt_tokens = max(1400, n_ctx - max_tokens - 512)
        return max(10_000, min(28_000, int(prompt_tokens * 3.0)))

    @staticmethod
    def _trim_middle(text: str, budget: int, label: str) -> str:
        if len(text) <= budget:
            return text
        marker = f"\n\n[{label} clipped locally to fit the model context.]\n\n"
        room = max(1000, budget - len(marker))
        head = max(500, int(room * 0.58))
        tail = max(500, room - head)
        return text[:head].rstrip() + marker + text[-tail:].lstrip()

    def _fit_messages_to_context(
        self,
        messages: list[dict[str, str]],
        max_tokens: int,
        budget_override: int | None = None,
    ) -> list[dict[str, str]]:
        budget = budget_override or self._context_char_budget(max_tokens)
        fitted = [dict(item) for item in messages]
        total = sum(len(item.get("content") or "") for item in fitted)
        if total <= budget:
            return fitted

        # First shrink older history while preserving the system prompt and newest request.
        for index in range(1, max(1, len(fitted) - 1)):
            content = fitted[index].get("content") or ""
            if len(content) > 900:
                fitted[index]["content"] = self._trim_middle(content, 900, "Older chat history")

        total = sum(len(item.get("content") or "") for item in fitted)
        if total <= budget:
            return fitted

        if fitted:
            last = fitted[-1]
            other_chars = sum(len(item.get("content") or "") for item in fitted[:-1])
            last_budget = max(5000, budget - other_chars)
            last["content"] = self._trim_middle(last.get("content") or "", last_budget, "Prompt context")

        total = sum(len(item.get("content") or "") for item in fitted)
        if total <= budget:
            return fitted

        system = fitted[0]
        other_chars = sum(len(item.get("content") or "") for item in fitted[1:])
        system_budget = max(3000, budget - other_chars)
        system["content"] = self._trim_middle(system.get("content") or "", system_budget, "System context")
        return fitted

    @property
    def model_path(self) -> Path:
        return Path(os.environ.get("CARAXES_MODEL_PATH", str(MODEL_FILE)))

    @property
    def openai_base_url(self) -> str | None:
        value = os.environ.get("CARAXES_OPENAI_BASE_URL", "").strip().rstrip("/")
        return value or None

    def status(self) -> dict[str, Any]:
        path = self.model_path
        base_url = self.openai_base_url
        return {
            "repo": MODEL_REPO,
            "backend": "openai-compatible" if base_url else "llama-cpp-python",
            "openai_base_url": base_url,
            "openai_model": os.environ.get("CARAXES_OPENAI_MODEL", path.name),
            "filename": path.name,
            "model_path": str(path),
            "model_exists": path.exists(),
            "model_size_gb": round(path.stat().st_size / 1024**3, 2) if path.exists() else None,
            "llama_cpp_installed": Llama is not None,
            "loaded": self._llm is not None,
            "busy": self._chat_lock.locked(),
            "last_error": self.last_error,
            "n_ctx": int(os.environ.get("CARAXES_N_CTX", "8192")),
            "n_threads": int(os.environ.get("CARAXES_N_THREADS", str(max(1, (os.cpu_count() or 4) - 1)))),
            "n_gpu_layers": int(os.environ.get("CARAXES_N_GPU_LAYERS", "0")),
        }

    def chat_openai_compatible(self, messages: list[dict[str, str]], temperature: float, max_tokens: int) -> str:
        base_url = self.openai_base_url
        if not base_url:
            raise RuntimeError("CARAXES_OPENAI_BASE_URL is not configured.")

        import httpx

        headers = {"Content-Type": "application/json"}
        api_key = os.environ.get("CARAXES_OPENAI_API_KEY")
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"

        fitted_messages = self._fit_messages_to_context(messages, max_tokens)
        payload = {
            "model": os.environ.get("CARAXES_OPENAI_MODEL", self.model_path.name),
            "messages": fitted_messages,
            "temperature": temperature,
            "max_tokens": max_tokens,
        }
        with httpx.Client(timeout=None) as client:
            response = client.post(f"{base_url}/chat/completions", headers=headers, json=payload)
            try:
                response.raise_for_status()
            except httpx.HTTPStatusError as exc:
                if response.status_code == 400 and "exceeds the available context size" in response.text:
                    retry_messages = self._fit_messages_to_context(messages, min(max_tokens, 1024), budget_override=14_000)
                    retry_payload = {**payload, "messages": retry_messages, "max_tokens": min(max_tokens, 1024)}
                    response = client.post(f"{base_url}/chat/completions", headers=headers, json=retry_payload)
                    try:
                        response.raise_for_status()
                    except httpx.HTTPStatusError as retry_exc:
                        detail = response.text[:1200].strip()
                        if detail:
                            raise RuntimeError(f"{retry_exc}. Backend response: {detail}") from retry_exc
                        raise
                    data = response.json()
                    message = data["choices"][0]["message"]
                    content = (message.get("content") or "").strip()
                    self.last_error = None
                    return content or "[The model returned an empty response after context clipping.]"
                detail = response.text[:1200].strip()
                if detail:
                    raise RuntimeError(f"{exc}. Backend response: {detail}") from exc
                raise
            data = response.json()
        message = data["choices"][0]["message"]
        content = (message.get("content") or "").strip()
        if content:
            self.last_error = None
            return content
        reasoning = (message.get("reasoning_content") or "").strip()
        if reasoning:
            self.last_error = None
            return (
                "The model used all generated tokens for internal reasoning and did not emit a final answer. "
                "Try increasing the token limit, or ask for a shorter direct answer."
            )
        self.last_error = None
        return "[The model returned an empty response.]"

    def load(self) -> Any:
        if Llama is None:
            raise RuntimeError("llama-cpp-python is not installed. Run: python -m pip install -r requirements-llm.txt")
        path = self.model_path
        if not path.exists():
            raise RuntimeError(f"Model file not found at {path}. Run scripts\\download_model.ps1 first.")
        with self._lock:
            if self._llm is None:
                self.last_error = None
                self._llm = Llama(
                    model_path=str(path),
                    n_ctx=int(os.environ.get("CARAXES_N_CTX", "8192")),
                    n_threads=int(os.environ.get("CARAXES_N_THREADS", str(max(1, (os.cpu_count() or 4) - 1)))),
                    n_gpu_layers=int(os.environ.get("CARAXES_N_GPU_LAYERS", "0")),
                    verbose=False,
                )
        return self._llm

    def chat(self, messages: list[dict[str, str]], temperature: float, max_tokens: int) -> str:
        with self._chat_lock:
            try:
                if self.openai_base_url:
                    return self.chat_openai_compatible(messages, temperature, max_tokens)
                llm = self.load()
                result = llm.create_chat_completion(
                    messages=messages,
                    temperature=temperature,
                    max_tokens=max_tokens,
                )
                self.last_error = None
                return result["choices"][0]["message"]["content"].strip()
            except Exception as exc:
                self.last_error = str(exc)
                raise


backfill_file_chunk_fts()
backfill_file_chunk_vectors()

runtime = ModelRuntime()
coder.init(
    runtime,
    DATA_DIR,
    ROOT,
    memory_provider=lambda: build_memory_context("__global__"),
    external_context_provider=lambda query: build_external_research_context(query),
)
app.include_router(coder.router)


def normalize_task_mode(mode: str | None) -> str:
    key = (mode or "auto").strip().lower().replace("-", "_").replace(" ", "_")
    return TASK_MODE_ALIASES.get(key, "auto")


def infer_task_mode(user_text: str, files: list[dict[str, Any]], requested_mode: str | None = None) -> str:
    mode = normalize_task_mode(requested_mode)
    if mode != "auto":
        return mode
    lower = user_text.lower()
    file_names = " ".join(item.get("original_name", "") for item in files).lower()
    text = f"{lower} {file_names}"
    if any(word in text for word in ("purple team", "purple-team", "emulation", "validate detection", "att&ck technique")):
        return "purple_team"
    if any(word in text for word in ("siem", "edr", "alert", "incident", "timeline", "forensic", "containment")):
        return "blue_team"
    if any(word in text for word in ("sigma", "yara", "detection rule", "false positive", "false negative")):
        return "detection_engineering"
    if any(word in text for word in ("red team", "attack path", "kill chain", "ttp", "adversary")):
        return "red_team_aware_defense"
    if any(word in text for word in ("cwe", "owasp", "vulnerability", "secure code", "security review")):
        return "secure_code"
    if any(word in text for word in ("review", "bug", "refactor", "test", "function", ".py", ".js", ".java", ".cpp")):
        return "code_review"
    if files and any((item.get("meta") or {}).get("chunk_count", 0) > 8 for item in files):
        return "large_file"
    return "coding"


def build_task_mode_context(mode: str) -> str:
    instruction = TASK_MODE_INSTRUCTIONS.get(mode)
    if not instruction:
        return ""
    return (
        "Task workflow instruction. Follow this shape unless the user explicitly asks for a different format.\n\n"
        + instruction
    )


def build_model_messages(
    chat_id: str,
    user_text: str,
    files: list[dict[str, Any]],
    mode: str = "auto",
    research_sources: list[dict[str, Any]] | None = None,
) -> list[dict[str, str]]:
    task_mode = infer_task_mode(user_text, files, mode)
    file_context = build_file_context(files, user_text)
    github_context = build_github_context(user_text, research_sources)
    web_context = build_web_context(user_text, research_sources)
    memory_context = build_memory_context(chat_id)
    task_context = build_task_mode_context(task_mode)
    system_parts = [SYSTEM_PROMPT]
    if memory_context:
        system_parts.append(memory_context)
    if task_context:
        system_parts.append(task_context)
    messages: list[dict[str, str]] = [{"role": "system", "content": "\n\n".join(system_parts)}]
    history_budget = MAX_HISTORY_CONTEXT_CHARS

    for message in recent_messages(chat_id):
        role = message["role"]
        content = message["content"]
        if role in {"user", "assistant"} and history_budget > 0:
            clipped = content[-history_budget:]
            messages.append({"role": role, "content": clipped})
            history_budget -= len(clipped)

    current_user_text = user_text
    context_blocks = [block for block in (file_context, github_context, web_context) if block]
    if context_blocks:
        context_text = "\n\n".join(context_blocks)
        current_user_text = (
            f"{context_text}\n\n"
            "User request:\n"
            f"{user_text}\n\n"
            "Answer the request directly. If this is code analysis, explain the purpose, flow, and important pieces. "
            "For large files, work from the included parts and the file profile; when more coverage is needed, name the next "
            "specific part numbers or search terms to inspect. Do not simply print attached or fetched source text."
        )
    messages.append({"role": "user", "content": current_user_text})
    return messages


def build_external_research_context(query: str) -> dict[str, Any]:
    sources: list[dict[str, Any]] = []
    blocks = [
        block
        for block in (
            build_github_context(query, sources),
            build_web_context(query, sources),
        )
        if block
    ]
    return {"context": "\n\n".join(blocks), "sources": sources}


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "home.html")


@app.get("/chat")
def chat_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/coder")
def coder_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "coder.html")


@app.get("/login")
def login_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "login.html")


@app.get("/activity")
def activity_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "activity.html")


@app.get("/api/status")
def api_status() -> dict[str, Any]:
    return {
        "app": APP_NAME,
        "model": runtime.status(),
    }


@app.get("/api/chats")
def list_chats(request: Request) -> list[dict[str, Any]]:
    user = accounts.require_user(request)
    with db() as conn:
        rows = conn.execute(
            "select * from chats where user_id = ? order by updated_at desc",
            (user["id"],),
        ).fetchall()
    return [row_to_dict(row) for row in rows]


@app.post("/api/chats")
def api_create_chat(payload: ChatCreate, request: Request) -> dict[str, Any]:
    user = accounts.require_user(request)
    return create_chat(payload.title, user["id"])


@app.get("/api/chats/{chat_id}")
def get_chat(chat_id: str, request: Request) -> dict[str, Any]:
    user = accounts.require_user(request)
    chat = get_chat_or_404(chat_id, user["id"])
    with db() as conn:
        rows = conn.execute(
            "select * from messages where chat_id = ? order by created_at asc",
            (chat_id,),
        ).fetchall()
    chat["messages"] = [row_to_dict(row) for row in rows]
    return chat


@app.delete("/api/chats/{chat_id}")
def delete_chat(chat_id: str, request: Request) -> dict[str, bool]:
    user = accounts.require_user(request)
    get_chat_or_404(chat_id, user["id"])
    with db() as conn:
        conn.execute("delete from messages where chat_id = ?", (chat_id,))
        conn.execute("delete from chats where id = ?", (chat_id,))
    return {"ok": True}


@app.post("/api/files")
def upload_files(files: list[UploadFile] = File(...)) -> list[dict[str, Any]]:
    saved = []
    for upload in files:
        saved.append(save_uploaded_file(upload))
    return saved


@app.post("/api/chats/{chat_id}/messages", response_model=ChatMessageOut)
def send_message(chat_id: str, payload: ChatMessageIn, request: Request) -> ChatMessageOut:
    user = accounts.require_user(request)
    chat = get_chat_or_404(chat_id, user["id"])
    accounts.log_activity(user, "chat.message", f'sent a message in chat "{chat.get("title") or "New chat"}"')
    content = payload.content.strip()
    if not content and not payload.attachment_ids:
        raise HTTPException(status_code=400, detail="Message or attachment required")

    files = fetch_files(payload.attachment_ids)
    if content:
        store_explicit_memories(content, f"chat:{chat_id}")
    attachments = [
        {
            "id": item["id"],
            "name": item["original_name"],
            "mime_type": item.get("mime_type"),
            "meta": item.get("meta"),
        }
        for item in files
    ]
    research_sources: list[dict[str, Any]] = []
    model_messages = build_model_messages(
        chat_id,
        content or "Please inspect the attached files.",
        files,
        payload.mode,
        research_sources,
    )
    user_message = insert_message(chat_id, "user", content or "[Attached files]", attachments)
    try:
        answer = runtime.chat(model_messages, payload.temperature, payload.max_tokens)
    except Exception as exc:
        if runtime.openai_base_url:
            answer = (
                "I could not get a response from the local model backend yet.\n\n"
                f"Reason: {exc}\n\n"
                f"The UI and file extraction are working. Check that the llama.cpp backend is still running at "
                f"`{runtime.openai_base_url}`."
            )
        else:
            answer = (
                "I could not run the local model yet.\n\n"
                f"Reason: {exc}\n\n"
                "The chat UI and file extraction are working, but the model needs the GGUF file and llama-cpp-python runtime. "
                "Run `scripts\\download_model.ps1`, then `python -m pip install -r requirements-llm.txt`, and restart the server."
            )
    assistant_attachments = [
        {
            "name": item.get("name") or item.get("url") or "source",
            "url": item.get("url"),
            "kind": item.get("kind", "web"),
            "fetched": item.get("fetched", False),
        }
        for item in research_sources[:WEB_MAX_RESULTS + 3]
        if item.get("url")
    ]
    assistant_message = insert_message(chat_id, "assistant", answer, assistant_attachments)
    update_chat_memory(chat_id, content or "[Attached files]", files, answer)
    return ChatMessageOut(
        chat_id=chat_id,
        user_message=user_message,
        assistant_message=assistant_message,
        model_status=runtime.status(),
        research_sources=research_sources,
    )
